import time
import pymysql
import json
import warnings
import openpyxl
from typing import List, Dict, Any, Optional
import re
from openpyxl.styles import PatternFill

warnings.filterwarnings("ignore")


########################################################################################################################
def save_to_excel(filename: str, query: str, real_results: List[Dict], videx_results: List[Dict],
                  comparison: Dict) -> None:
    # 创建Excel工作簿
    wb = openpyxl.Workbook()

    # 工作表1: 真实数据库EXPLAIN结果
    sheet1 = wb.active
    sheet1.title = "真实数据库"
    _add_explain_results(sheet1, query, "真实数据库", real_results)

    # 工作表2: Videx数据库EXPLAIN结果
    sheet2 = wb.create_sheet(title="Videx数据库")
    _add_explain_results(sheet2, query, "Videx数据库", videx_results)

    # 工作表3: 比较结果
    sheet3 = wb.create_sheet(title="比较结果")
    _add_comparison(sheet3, query, comparison)

    # 保存文件
    wb.save(filename)


def _add_explain_results(sheet, query: str, db_name: str, results: List[Dict]) -> None:
    """将EXPLAIN结果添加到工作表"""
    # 添加标题
    sheet.append([f"{db_name} - EXPLAIN 结果"])
    sheet.append(["SQL查询:", query])
    sheet.append([])

    # 添加表头
    if not results:
        sheet.append(["无可用结果"])
        return

    # 获取所有列名
    columns = list(results[0].keys())
    sheet.append([""] + columns)

    # 添加数据行
    for idx, row in enumerate(results):
        # 添加行号和行数据
        row_data = [idx + 1] + [row.get(col, "") for col in columns]
        sheet.append(row_data)


def _add_comparison(sheet, query: str, comparison: Dict) -> None:
    """将比较结果添加到工作表"""
    # 添加标题
    sheet.append(["EXPLAIN 比较结果"])
    sheet.append(["SQL查询:", query])
    sheet.append(["得分:", comparison.get("score", 0)])
    sheet.append(["摘要:", comparison.get("msg", "")])
    sheet.append([])

    # 添加差异标题
    sheet.append(["差异详情"])
    sheet.append(["字段", "期望值", "实际值"])

    # 添加差异数据
    differences = comparison.get("differences", {})
    for field, diff_list in differences.items():
        sheet.append([
            field,
            diff_list.get("expected", ""),
            diff_list.get("actual", "")
        ])

    # 突出显示差异行
    _highlight_differences(sheet)


def _highlight_differences(sheet):
    """高亮显示差异行"""
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 从第8行开始是差异数据（标题行+1为差异行）
    for row_idx in range(8, sheet.max_row + 1):
        for col in range(1, 5):
            cell = sheet.cell(row=row_idx, column=col)
            cell.fill = red_fill


########################################################################################################################
def safe_convert(value):
    """将值转换为Excel可接受的格式"""
    if value is None:
        return ""
    elif isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    elif isinstance(value, float):
        return round(value, 2)
    return value


def save_results_to_excel(query_num, sql, real_explain, videx_explain, comparison_result, filename):
    # 创建基础工作簿
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        # 删除默认创建的工作表
        if wb.sheetnames and wb.sheetnames[0] == 'Sheet':
            del wb['Sheet']

    # 1. 保存原始解释结果
    sheet_name_real = f"Query{query_num}_Real"
    if sheet_name_real in wb.sheetnames:
        del wb[sheet_name_real]
    sheet1 = wb.create_sheet(title=sheet_name_real)
    _add_explain_to_sheet(sheet1, sql, real_explain)

    # 2. 保存VIDEX解释结果
    sheet_name_videx = f"Query{query_num}_Videx"
    if sheet_name_videx in wb.sheetnames:
        del wb[sheet_name_videx]
    sheet2 = wb.create_sheet(title=sheet_name_videx)
    _add_explain_to_sheet(sheet2, sql, videx_explain)

    # 3. 保存比较结果
    sheet_name_comp = f"Query{query_num}_Comparison"
    if sheet_name_comp in wb.sheetnames:
        del wb[sheet_name_comp]
    sheet3 = wb.create_sheet(title=sheet_name_comp)
    _add_comparison_to_sheet(sheet3, sql, comparison_result)

    # 4. 保存差异详情
    sheet_name_diff = f"Query{query_num}_DiffDetails"
    if sheet_name_diff in wb.sheetnames:
        del wb[sheet_name_diff]
    sheet4 = wb.create_sheet(title=sheet_name_diff)
    _add_diff_details_to_sheet(sheet4, sql, comparison_result['diff'])

    wb.save(filename)
    return wb


def _add_explain_to_sheet(sheet, sql, explain_data):
    """将解释结果添加到工作表"""
    # 添加SQL查询
    sheet.append(["SQL查询:", sql])

    if not explain_data:
        sheet.append(["无可用解释结果"])
        return

    # 添加表头
    headers = ["字段路径", "字段值"]
    sheet.append(headers)

    # 添加解释数据
    flattened = flatten_explain_data(explain_data)
    for path, value in flattened.items():
        safe_value = safe_convert(value)
        sheet.append([path, safe_value])


def _add_comparison_to_sheet(sheet, sql, comparison_result):
    """将比较结果添加到工作表"""
    # 添加SQL查询
    sheet.append(["SQL查询:", sql])
    sheet.append(["比较得分:", comparison_result['score']])
    sheet.append(["差异摘要:", comparison_result.get('msg', '无显著差异')])
    sheet.append([])  # 空行分隔

    if not comparison_result['diff']:
        sheet.append(["无显著差异"])
        return

    # 添加表头
    headers = ["执行计划项", "字段", "期望值", "实际值"]
    sheet.append(headers)

    # 添加差异数据
    for item_idx, item_diffs in comparison_result['diff'].items():
        for field, diff_info in item_diffs.items():
            expected = safe_convert(diff_info.get("expected", ""))
            actual = safe_convert(diff_info.get("actual", ""))

            sheet.append([
                f"项目 #{item_idx}",
                field,
                expected,
                actual
            ])


def _add_diff_details_to_sheet(sheet, sql, diff_data):
    """将详细差异信息添加到工作表"""
    # 添加SQL查询
    sheet.append(["SQL查询:", sql])
    sheet.append(["详细差异分解:"])

    if not diff_data:
        sheet.append(["无详细差异信息"])
        return

    row_idx = 3  # 从第3行开始

    for item_idx, item_diffs in diff_data.items():
        sheet.cell(row=row_idx, column=1, value=f"执行计划项 #{item_idx}")
        sheet.cell(row=row_idx, column=1).font = openpyxl.styles.Font(bold=True)
        row_idx += 1

        for field, diff_info in item_diffs.items():
            # 添加主字段差异
            sheet.cell(row=row_idx, column=1, value="字段:")
            sheet.cell(row=row_idx, column=2, value=field)
            row_idx += 1

            # 添加期望值和实际值
            sheet.cell(row=row_idx, column=1, value="期望值:")
            sheet.cell(row=row_idx, column=2, value=safe_convert(diff_info.get("expected", "")))
            row_idx += 1

            sheet.cell(row=row_idx, column=1, value="实际值:")
            sheet.cell(row=row_idx, column=2, value=safe_convert(diff_info.get("actual", "")))
            row_idx += 1

            # 添加详细差异
            details = diff_info.get("details", {})
            if details:
                sheet.cell(row=row_idx, column=1, value="详细差异:")
                row_idx += 1

                for sub_key, sub_diff in details.items():
                    sheet.cell(row=row_idx, column=2, value=sub_key)
                    row_idx += 1

                    sheet.cell(row=row_idx, column=3, value="期望:")
                    sheet.cell(row=row_idx, column=4, value=safe_convert(sub_diff.get("expected", "")))
                    row_idx += 1

                    sheet.cell(row=row_idx, column=3, value="实际:")
                    sheet.cell(row=row_idx, column=4, value=safe_convert(sub_diff.get("actual", "")))
                    row_idx += 1
            else:
                sheet.cell(row=row_idx, column=1, value="无额外详细信息")
                row_idx += 1

            row_idx += 1  # 项目间空行


def flatten_explain_data(data, parent_key='', sep='.'):
    """展平解释数据结构，方便导出到Excel"""
    items = {}
    if isinstance(data, dict):
        for k, v in data.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.update(flatten_explain_data(v, new_key, sep=sep))
            elif isinstance(v, list):
                # 将列表转换为索引项目
                for i, item in enumerate(v):
                    items[f"{new_key}[{i}]"] = safe_convert(item)
                # 同时保存整个列表作为JSON字符串
                items[new_key] = safe_convert(v)
            else:
                items[new_key] = safe_convert(v)
    return items


########################################################################################################################
def compare_explain(expect: Dict, actual: Dict) -> Dict:
    """比较两个EXPLAIN结果"""
    # 初始化比较结果
    result = {
        "matched": True,
        "differences": {}
    }

    # 比较所有公共键
    all_keys = set(expect.keys()) | set(actual.keys())

    for key in all_keys:
        # 处理值不存在的情况
        exp_val = expect.get(key)
        act_val = actual.get(key)

        # 如果是列表类型，比较每个元素
        if isinstance(exp_val, list) and isinstance(act_val, list):
            # 比较列表长度
            if len(exp_val) != len(act_val):
                result["matched"] = False
                result["differences"][key] = {
                    "expected": exp_val,
                    "actual": act_val,
                    "message": f"列表长度不同: {len(exp_val)} vs {len(act_val)}"
                }
                continue

            # 比较列表中的每个元素
            for i, (exp_item, act_item) in enumerate(zip(exp_val, act_val)):
                if exp_item != act_item:
                    result["matched"] = False
                    result["differences"].setdefault(key, []).append({
                        "index": i,
                        "expected": exp_item,
                        "actual": act_item
                    })

        # 对于非列表类型的值
        elif exp_val != act_val:
            result["matched"] = False
            result["differences"][key] = {
                "expected": exp_val,
                "actual": act_val
            }

    return result


########################################################################################################################
def compare_explain_with_json(expect: List[Dict], actual: List[Dict]):
    # 长度检查
    if len(expect) != len(actual):
        return {
            "score": 0.0,
            "msg": "解释项的长度不匹配",
            "diff": {}
        }

    def extract_cost_info(plan):
        """从执行计划中提取成本信息"""
        costs = {}
        # 提取查询块级别的成本
        if 'query_block' in plan and 'cost_info' in plan['query_block']:
            costs.update(plan['query_block']['cost_info'])

        # 提取表格级别的成本
        if 'query_block' in plan and 'table' in plan['query_block']:
            table = plan['query_block']['table']
            if 'cost_info' in table:
                costs.update({f"table_{k}": v for k, v in table['cost_info'].items()})
        return costs

    def compare_keys(explain_a, explain_b, key):
        """比较特定键的值"""
        # 处理成本信息
        if key == 'cost_info':
            cost_a = extract_cost_info(explain_a)
            cost_b = extract_cost_info(explain_b)

            diffs = {}
            for k in set(cost_a.keys()) | set(cost_b.keys()):
                val_a = cost_a.get(k, 0)
                val_b = cost_b.get(k, 0)

                try:
                    num_a = float(val_a) if not isinstance(val_a, float) else val_a
                    num_b = float(val_b) if not isinstance(val_b, float) else val_b
                except (TypeError, ValueError):
                    # 非数值类型直接比较
                    if val_a != val_b:
                        diffs[k] = {"expected": val_a, "actual": val_b}
                    continue

                # 忽略都为0的情况
                if num_a == 0 and num_b == 0:
                    continue

                # 数值比较逻辑
                abs_diff = abs(num_a - num_b)
                # 处理零值情况
                base_value = max(num_a, num_b)
                rel_diff = abs_diff / base_value if base_value > 0 else 0

                # 宽松比较策略: 15%的相对差异或小于5的绝对差异
                if abs_diff > max(5, 0.15 * base_value):
                    diffs[k] = {"expected": num_a, "actual": num_b}

            return len(diffs) == 0, cost_a, cost_b, diffs

        # 处理引用字段
        if key == 'ref':
            ref_a = explain_a.get('query_block', {}).get('table', {}).get('ref', [])
            ref_b = explain_b.get('query_block', {}).get('table', {}).get('ref', [])

            # 处理空值情况
            if not ref_a and not ref_b:
                return True, ref_a, ref_b, {}

            # 比较引用值，忽略数据库名
            refs_match = True
            diff_details = {}
            for i, (r_a, r_b) in enumerate(zip(ref_a, ref_b)):
                # 去除数据库名
                if isinstance(r_a, str):
                    r_a = r_a.split('.')[-1]
                if isinstance(r_b, str):
                    r_b = r_b.split('.')[-1]

                if r_a != r_b:
                    refs_match = False
                    diff_details[f"ref[{i}]"] = {
                        "expected": r_a,
                        "actual": r_b
                    }

            return refs_match, ref_a, ref_b, diff_details

        # 处理扫描行数
        if key == 'rows':
            rows_a = explain_a.get('query_block', {}).get('table', {}).get('rows_examined_per_scan') or \
                     explain_a.get('query_block', {}).get('table', {}).get('rows', 0)
            rows_b = explain_b.get('query_block', {}).get('table', {}).get('rows_examined_per_scan') or \
                     explain_b.get('query_block', {}).get('table', {}).get('rows', 0)

            try:
                num_a = float(rows_a)
                num_b = float(rows_b)
                diff_val = abs(num_a - num_b)

                # 行数比较逻辑
                if min(num_a, num_b) < 10:
                    # 小表: 允许1行差异
                    match = diff_val <= 1
                else:
                    # 大表: 允许5%差异或最多50行差异
                    rel_diff = diff_val / max(num_a, num_b)
                    match = (rel_diff <= 0.05) or (diff_val <= 50)

                return match, num_a, num_b, {}
            except (TypeError, ValueError):
                return rows_a == rows_b, rows_a, rows_b, {}

        # 处理附加条件
        if key in ['attached_condition', 'index_condition']:
            val_a = explain_a.get('query_block', {}).get('table', {}).get(key, '')
            val_b = explain_b.get('query_block', {}).get('table', {}).get(key, '')

            # 标准化字符串比较
            clean_a = re.sub(r'`\w+`\.', '', str(val_a)).strip()  # 移除数据库名
            clean_b = re.sub(r'`\w+`\.', '', str(val_b)).strip()

            # 忽略大小写和空格差异
            match = clean_a.lower() == clean_b.lower()
            return match, val_a, val_b, {}

        # 处理Extra字段
        if key == 'Extra':
            val_a = explain_a.get('query_block', {}).get('table', {}).get('Extra', '')
            val_b = explain_b.get('query_block', {}).get('table', {}).get('Extra', '')

            # 标准化字符串比较
            clean_a = re.sub(r'[\s;]+', ' ', str(val_a)).strip().lower()
            clean_b = re.sub(r'[\s;]+', ' ', str(val_b)).strip().lower()

            match = clean_a == clean_b
            return match, val_a, val_b, {}

        # 默认比较逻辑
        val_a = explain_a.get('query_block', {}).get('table', {}).get(key)
        val_b = explain_b.get('query_block', {}).get('table', {}).get(key)
        return val_a == val_b, val_a, val_b, {}

    # 关键字段列表
    keys_to_compare = [
        'table',
        'select_type',
        'type',
        'ref',
        'key',
        'key_len',
        'possible_keys',
        'rows',
        'Extra',
        'attached_condition',
        'index_condition',
        'cost_info'
    ]

    matched_count = 0
    all_diffs = {}
    first_mismatch_msg = ""

    for idx, (e_item, a_item) in enumerate(zip(expect, actual)):
        item_diffs = {}
        item_match = True

        for key in keys_to_compare:
            # 获取比较结果
            is_match, e_val, a_val, key_diffs = compare_keys(e_item, a_item, key)

            if not is_match:
                item_match = False
                # 确保值可以安全序列化
                if isinstance(e_val, (dict, list)):
                    e_val = json.dumps(e_val, ensure_ascii=False)
                if isinstance(a_val, (dict, list)):
                    a_val = json.dumps(a_val, ensure_ascii=False)

                item_diffs[key] = {
                    "expected": e_val,
                    "actual": a_val,
                    "details": key_diffs
                }

                # 记录第一条差异信息
                if not first_mismatch_msg:
                    # 限制消息长度
                    e_val_str = str(e_val)[:50] + "..." if len(str(e_val)) > 50 else str(e_val)
                    a_val_str = str(a_val)[:50] + "..." if len(str(a_val)) > 50 else str(a_val)
                    first_mismatch_msg = f"执行计划项 #{idx}, 字段={key}, 期望={e_val_str}, 实际={a_val_str}"

        if item_match:
            matched_count += 1
        else:
            all_diffs[idx] = item_diffs

    score = matched_count / len(expect) if expect else 0.0

    return {
        "score": score,
        "msg": first_mismatch_msg,
        "diff": all_diffs
    }


########################################################################################################################
def get_storage_engine(db_config, table_name):
    global connection
    try:
        connection = pymysql.connect(**db_config)
        cursor = connection.cursor()
        query = f"SHOW TABLE STATUS LIKE '{table_name}'"
        cursor.execute(query)
        result = cursor.fetchone()
        if result:
            return result[1]  # 存储引擎信息在结果的第二列
        return None
    except pymysql.Error as e:
        print(f"获取存储引擎信息时出错: {e}")
    finally:
        if 'connection' in locals() and connection:
            connection.close()


########################################################################################################################
def analyze_query_with_index(db_config: Dict, table_name: str, sql: str, index_name: Optional[str] = None) -> \
        Optional[Dict]:
    """分析查询性能（可选择是否使用特定索引）"""
    if table_name not in table_configs:
        print(f"未找到表 {table_name} 的配置")
        return None

    with IndexManager(db_config) as manager:
        # 如果有指定索引，先删除再创建
        if index_name:
            index_info = table_configs[table_name]['indexes'].get(index_name)
            if not index_info:
                print(f"表 {table_name} 中未找到索引 {index_name}")
                return None

            # 先删除索引（如果存在）
            manager.drop_index(table_name, index_name)

            # 创建索引
            if not manager.create_index(table_name, index_name, index_info['columns']):
                return None

        # 执行EXPLAIN分析
        explain_result = manager.execute_explain(table_name, sql)

        return explain_result


########################################################################################################################
def compare_cost_info(expect: Dict, actual: Dict) -> Dict:
    """
    专门比较两个执行计划的cost_info差异
    返回包含详细比较结果的字典
    """

    def extract_cost_details(plan: Dict) -> Dict:
        """从执行计划中提取所有成本相关信息"""
        costs = {}

        # 提取查询块级别的成本
        if 'query_block' in plan and 'cost_info' in plan['query_block']:
            costs.update(plan['query_block']['cost_info'])

        # 提取表格级别的成本
        if 'query_block' in plan and 'table' in plan['query_block']:
            table = plan['query_block']['table']
            if 'cost_info' in table:
                costs.update({f"table_{k}": v for k, v in table['cost_info'].items()})

        return costs

    # 提取成本信息
    expected_costs = extract_cost_details(expect)
    actual_costs = extract_cost_details(actual)

    # 比较结果
    comparison = {
        'cost_fields': sorted(set(expected_costs.keys()) | set(actual_costs.keys())),
        'details': {},
        'significant_diffs': []
    }

    # 数值比较阈值
    ABS_DIFF_THRESHOLD = 3
    REL_DIFF_THRESHOLD = 0.15  # 15%

    for field in comparison['cost_fields']:
        expected_val = expected_costs.get(field, 0)
        actual_val = actual_costs.get(field, 0)

        try:
            # 尝试转换为数值比较
            expected_num = float(expected_val) if not isinstance(expected_val, (int, float)) else expected_val
            actual_num = float(actual_val) if not isinstance(actual_val, (int, float)) else actual_val

            abs_diff = abs(expected_num - actual_num)
            rel_diff = abs_diff / max(expected_num, actual_num) if max(expected_num, actual_num) > 0 else 0

            is_significant = (abs_diff > ABS_DIFF_THRESHOLD and
                              rel_diff > REL_DIFF_THRESHOLD and
                              not (expected_num == 0 and actual_num == 0))

            comparison['details'][field] = {
                'expected': expected_val,
                'actual': actual_val,
                'abs_diff': f"{abs_diff:.2}",
                'rel_diff': f"{rel_diff:.2%}",
                'is_significant': is_significant
            }

            if is_significant:
                comparison['significant_diffs'].append({
                    'field': field,
                    'expected': expected_val,
                    'actual': actual_val,
                    'abs_diff': f"{abs_diff:.2}",
                    'rel_diff': f"{rel_diff:.2%}"
                })

        except (ValueError, TypeError):
            # 非数值类型直接比较
            comparison['details'][field] = {
                'expected': expected_val,
                'actual': actual_val,
                'is_same': expected_val == actual_val
            }

    return comparison


########################################################################################################################
def format_cost_comparison(comparison: Dict) -> str:
    """格式化成本比较结果为可读字符串"""
    lines = ["成本信息比较结果:"]
    lines.append("-" * 60)

    if not comparison['significant_diffs']:
        lines.append("无显著差异")
    else:
        lines.append("显著差异字段:")
        for diff in comparison['significant_diffs']:
            lines.append(
                f"  {diff['field']}: 期望={diff['expected']}, 实际={diff['actual']} "
                f"(绝对差={diff['abs_diff']}, 相对差={diff['rel_diff']})"
            )

    lines.append("\n所有成本字段详情:")
    for field in comparison['cost_fields']:
        detail = comparison['details'][field]
        if 'is_same' in detail:
            status = "相同" if detail['is_same'] else "不同"
            lines.append(f"  {field}: {status} (期望={detail['expected']}, 实际={detail['actual']})")
        else:
            sig_flag = "!" if detail['is_significant'] else ""
            lines.append(
                f"  {field}{sig_flag}: 期望={detail['expected']}, 实际={detail['actual']} "
                f"(绝对差={detail['abs_diff']}, 相对差={detail['rel_diff']})"
            )

    return "\n".join(lines)


########################################################################################################################
########################################################################################################################
class IndexManager:
    def __init__(self, db_config: Dict):
        self.db_config = db_config
        self.connection = None

    def __enter__(self):
        self.connection = pymysql.connect(**self.db_config)
        self.connection.autocommit(True)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.connection:
            self.connection.close()

    def drop_index(self, table_name: str, index_name: str) -> bool:
        """删除指定索引（如果存在）"""
        try:
            with self.connection.cursor() as cursor:
                cursor.execute(f"SHOW INDEX FROM {table_name} WHERE Key_name = %s", (index_name,))
                if cursor.fetchone():
                    drop_sql = f"ALTER TABLE {table_name} DROP INDEX {index_name}"
                    cursor.execute(drop_sql)
                    return True
                print(f"索引 {index_name} 不存在，无需删除")
                return False
        except pymysql.Error as e:
            print(f"删除索引 {index_name} 时出错: {e}")
            return False

    def create_index(self, table_name: str, index_name: str, columns: str) -> bool:
        """创建新索引（如果不存在）"""
        try:
            with self.connection.cursor() as cursor:
                cursor.execute(f"SHOW INDEX FROM {table_name} WHERE Key_name = %s", (index_name,))
                if not cursor.fetchone():
                    create_sql = f"ALTER TABLE {table_name} ADD INDEX {index_name} ({columns})"
                    cursor.execute(create_sql)
                    print(f"已创建索引 {index_name} ({columns})")
                    return True
                print(f"索引 {index_name} 已存在，无需创建")
                return False
        except pymysql.Error as e:
            print(f"创建索引 {index_name} 时出错: {e}")
            return False

    def execute_explain_with_json(self, sql: str) -> Optional[Dict]:
        """执行EXPLAIN分析"""
        try:
            with self.connection.cursor() as cursor:
                explain_sql = f"EXPLAIN FORMAT=JSON {sql}"
                cursor.execute(explain_sql)
                result = cursor.fetchone()
                return json.loads(result[0]) if result else None
        except pymysql.Error as e:
            print(f"执行EXPLAIN时出错: {e}")
            return None

    def execute_explain(self, sql: str) -> Optional[Dict]:
        """执行EXPLAIN分析并返回结构化的结果"""
        try:
            with self.connection.cursor() as cursor:
                explain_sql = f"EXPLAIN {sql}"
                cursor.execute(explain_sql)
                columns = [col[0] for col in cursor.description]
                explain_rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

                if len(explain_rows) == 1:
                    return explain_rows[0]

                return {
                    "id": [row["id"] for row in explain_rows],
                    "select_type": [row["select_type"] for row in explain_rows],
                    "table": [row["table"] for row in explain_rows],
                    "partitions": [row["partitions"] for row in explain_rows],
                    "type": [row["type"] for row in explain_rows],
                    "possible_keys": [row["possible_keys"] for row in explain_rows],
                    "key": [row["key"] for row in explain_rows],
                    "key_len": [row["key_len"] for row in explain_rows],
                    "ref": [row["ref"] for row in explain_rows],
                    "rows": [row["rows"] for row in explain_rows],
                    "filtered": [row["filtered"] for row in explain_rows],
                    "Extra": [row["Extra"] for row in explain_rows]
                }
        except pymysql.Error as e:
            print(f"执行EXPLAIN时出错: {e}")
            return None


def initialize_indexes(manager: IndexManager, table_config: dict, table_name: str):
    """初始化索引：删除已存在的索引并创建新索引"""
    indexes = table_config['indexes']
    for index_name in indexes.keys():
        manager.drop_index(table_name, index_name)
    for index_name, index_info in indexes.items():
        manager.create_index(table_name, index_name, index_info['columns'])


def analyze_query_with_json(manager: IndexManager, sql: str) -> Optional[Dict]:
    """分析查询性能（不处理索引）"""
    return manager.execute_explain_with_json(sql)


def analyze_query(manager: IndexManager, sql: str) -> Optional[Dict]:
    """分析查询性能（不处理索引）"""
    return manager.execute_explain(sql)


########################################################################################################################
########################################################################################################################
# 数据库连接信息
# real_db_config = {
#     'host': 'rm-uf6pyrv408i5f0gap.mysql.rds.aliyuncs.com',
#     'port': 3306,
#     'user': 'onedba',
#     'password': 'S9dKSCsdJm(mKd2',
#     'database': 'du_trade_timeout_db_3'
# }

real_db_config = {
    'host': '127.0.0.1',
    'port': 13308,
    'user': 'videx',
    'password': 'password',
    'database': 'tpch_tiny'
}

videx_db_config = {
    'host': '127.0.0.1',
    'port': 13308,
    'user': 'videx',
    'password': 'password',
    'database': 'videx_1'
}

# 配置数据结构 - 可以扩展为支持多个表
table_configs = {
    't_users': {
        'indexes': {
            'idx_name': {'columns': 'name', 'type': 'single'},
            'idx_active_city_age': {'columns': 'is_active, city, age', 'type': 'composite'}
        },
        'queries': [
            "SELECT * FROM t_users WHERE name = '张%'",
            "SELECT * FROM t_users WHERE is_active = 1 AND city LIKE '北%' AND age > 20"
            # "SELECT city,AVG(age) AS avg_age,COUNT(*) AS active_users FROM t_users WHERE is_active = 1 \
            #     GROUP BY city HAVING AVG(age) < 50 AND COUNT(*) > 20 UNION ALL \
            #     SELECT city, NULL AS avg_age, NULL AS active_users \
            #     FROM (SELECT city, created_at, ROW_NUMBER() OVER (PARTITION BY city ORDER BY created_at DESC) AS rn \
            #     FROM t_users WHERE is_active = 0) AS inactive_users WHERE rn = 1 \
            #     ORDER BY CASE WHEN avg_age IS NOT NULL THEN 1 ELSE 2 END, active_users DESC;"
        ]
    },

    # 'timeout_record_31': {
    #     'indexes': {
    #         'idx_source_system': {'columns': 'source_system', 'type': 'single'},
    #         'idx_biz_id_result': {'columns': 'biz_id, result', 'type': 'composite'}
    #     },
    #     'queries': [
    #         "SELECT * FROM timeout_record_31;"
    #     ]
    # }
}


########################################################################################################################
########################################################################################################################
def convert_to_list(opt_dict: Optional[dict]) -> List[dict]:
    return [opt_dict] if opt_dict else []


if __name__ == "__main__":

    # table_name = 'timeout_record_31'
    table_name = 't_users'
    # table_name = 'orders'
    results_filename = "result/explain_results_with_json.xlsx"

    # 创建初始Excel文件
    wb = openpyxl.Workbook()
    wb.save(results_filename)

    # 获取存储引擎信息
    real_engine = get_storage_engine(real_db_config, table_name)
    videx_engine = get_storage_engine(videx_db_config, table_name)
    print(f"真实数据库 {real_db_config['database']} 中 {table_name} 表使用的存储引擎: {real_engine}")
    print(f"虚拟数据库 {videx_db_config['database']} 中 {table_name} 表使用的存储引擎: {videx_engine}")

    # 初始化数据库连接和索引
    with IndexManager(real_db_config) as real_manager, IndexManager(videx_db_config) as videx_manager:
        # 初始化索引（删除旧索引并创建新索引）
        table_config = table_configs[table_name]
        initialize_indexes(real_manager, table_config, table_name)
        initialize_indexes(videx_manager, table_config, table_name)

        # 遍历所有查询
        for i, sql_query in enumerate(table_configs[table_name]['queries']):
            print(f"\n测试查询 {i + 1}: {sql_query}")

            # 在真实数据库中执行EXPLAIN
            start_time = time.time()
            real_explain = analyze_query_with_json(real_manager, sql_query)
            end_time = time.time()
            print("真实数据库执行时间: ", f"{(end_time - start_time) * 1000:.2f}ms")
            print("原始数据库中执行 explain format=json 结果如下：\n", real_explain, "\n")

            # 在虚拟数据库中执行EXPLAIN
            start_time = time.time()
            videx_explain = analyze_query_with_json(videx_manager, sql_query)
            end_time = time.time()
            print("虚拟数据库执行时间: ", f"{(end_time - start_time) * 1000:.2f}ms")
            print("Videx 数据库中执行 explain format=json 结果如下：\n", videx_explain, "\n")

            if real_explain and videx_explain:
                comparison_result = compare_explain_with_json([real_explain], [videx_explain])
                # 新增成本详细比较
                cost_comparison = compare_cost_info(real_explain, videx_explain)
                print(format_cost_comparison(cost_comparison))

                print("比较结果:")
                print(f"得分: {comparison_result['score']}")
                print(f"简要消息: {comparison_result['msg']}")
                print(f"详细差异: {comparison_result['diff']}")

                # # 保存结果到Excel
                # wb = save_results_to_excel(
                #     i + 1,
                #     sql_query,
                #     real_explain,
                #     videx_explain,
                #     comparison_result,
                #     results_filename
                # )
            else:
                print(f"跳过查询 {i + 1}，获取解释计划失败")

            print("*" * 60)
            ########################################################################################################################
            results_filename = 'result/explain_results.xlsx'

            # 在真实数据库中执行EXPLAIN
            start_time = time.time()
            real_explain = analyze_query(real_manager, sql_query)
            end_time = time.time()
            print("原始数据库中执行 explain 耗时：", f"{(end_time - start_time) * 1000:.2f}ms")
            print("原始数据库中执行 explain 结果如下：\n", real_explain, "\n")

            # 在VIDEX虚拟数据库中执行EXPLAIN
            start_time = time.time()
            videx_explain = analyze_query(videx_manager, sql_query)
            end_time = time.time()
            print("Videx 数据库中执行 explain 耗时：", f"{(end_time - start_time) * 1000:.2f}ms")
            print("Videx 数据库中执行 explain 结果如下：\n", videx_explain, "\n")

            if real_explain and videx_explain:
                # 调用compare_explain函数比较结果
                comparison_result = compare_explain(real_explain, videx_explain)
                print("匹配:", comparison_result["matched"])
                print("差异:", json.dumps(comparison_result["differences"], indent=2))

                # save_to_excel(results_filename, sql_query, convert_to_list(real_explain),
                #               convert_to_list(videx_explain), comparison_result)
                # print(f"结果已保存到 {results_filename}\n")
